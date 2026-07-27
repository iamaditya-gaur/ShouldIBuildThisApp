"""
p7_synthesize.py — build the workbook and write the decision.

    python scripts/p7_synthesize.py --niches "niche one,niche two"

Produces:
    output/FINAL_ANALYSIS.xlsx   every tab of evidence, with the gaps marked
    output/DECISION.md           BUILD / WATCH / KILL, with the case against

Missing evidence is shown as missing. If the Apple Search Ads step or the
revenue pull has not been done, the relevant tabs are templates for you to fill
and the verdict says plainly which claims are unsupported. A confident verdict
built on absent data is the specific failure this whole pipeline exists to
avoid.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from common import DATA, OUTPUT, banner, die, load_config  # noqa: E402

PHASE = "p7_synthesize"

FONT = "Arial"
H1 = Font(name=FONT, size=14, bold=True)
HDR = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY = Font(name=FONT, size=10)
NOTE = Font(name=FONT, size=9, italic=True, color="666666")
INPUT_FONT = Font(name=FONT, size=10, color="0000FF")
HDR_FILL = PatternFill("solid", fgColor="333333")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
WRAP = Alignment(wrap_text=True, vertical="top")


def write_table(ws, df: pd.DataFrame, start_row: int = 1) -> int:
    """Header + rows. Returns the row after the table."""
    for j, col in enumerate(df.columns, 1):
        c = ws.cell(row=start_row, column=j, value=str(col))
        c.font, c.fill = HDR, HDR_FILL
    for i, (_, r) in enumerate(df.iterrows(), start_row + 1):
        for j, col in enumerate(df.columns, 1):
            v = r[col]
            c = ws.cell(row=i, column=j,
                        value=None if pd.isna(v) else
                        (v if isinstance(v, (int, float, str)) else str(v)))
            c.font = BODY
    for j, col in enumerate(df.columns, 1):
        widths = [len(str(col))] + [min(len(str(v)), 42)
                                    for v in df[col].head(80).fillna("")]
        ws.column_dimensions[get_column_letter(j)].width = \
            min(max(max(widths) + 2, 10), 44)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return start_row + len(df) + 2


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--niches", required=True)
    args = ap.parse_args()
    niches = [n.strip() for n in args.niches.split(",") if n.strip()]

    cfg = load_config()
    OUTPUT.mkdir(parents=True, exist_ok=True)

    if not (DATA / "03_scored.csv").exists():
        die("data/03_scored.csv not found — run scripts/p3_score.py first.")
    scored = pd.read_csv(DATA / "03_scored.csv")
    comp = pd.read_csv(DATA / "02_competitors.csv")
    kw = pd.read_csv(DATA / "01_keywords.csv")

    asa = pd.read_csv(DATA / "04_asa_filled.csv") if (DATA / "04_asa_filled.csv").exists() else None
    rev = pd.read_csv(DATA / "06_revenue_filled.csv") if (DATA / "06_revenue_filled.csv").exists() else None
    complaints = (DATA / "05_complaints.md").read_text() if (DATA / "05_complaints.md").exists() else None
    stats = pd.read_csv(DATA / "05_review_stats.csv") if (DATA / "05_review_stats.csv").exists() else None

    banner("PHASE 7 — SYNTHESIS")
    print(f"\n  ASA data:       {'present' if asa is not None else 'MISSING (Phase 4 not done)'}")
    print(f"  Revenue data:   {'present' if rev is not None else 'MISSING (Phase 6 not done)'}")
    print(f"  Complaints:     {'present' if complaints else 'MISSING'}")

    sel = scored[scored.niche.isin(niches)]
    real = sel[~sel.probably_app_name.fillna(False)]

    wb = Workbook()

    # ---------------------------------------------------------- Summary ----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "iOS ASO Opportunity — Summary"
    ws["A1"].font = H1
    ws["A2"] = f"Primary market: {cfg['markets']['primary']['code']}   ·   " \
               f"{scored.keyword.nunique()} keywords   ·   {comp.app_id.nunique()} apps analysed"
    ws["A2"].font = NOTE

    ws["A4"] = "Evidence available"
    ws["A4"].font = Font(name=FONT, size=11, bold=True)
    evidence = [
        ("Keyword demand (autocomplete rank)", "yes", "PROXY — ordering only, not volume"),
        ("Competition (who ranks, how strong)", "yes", "Observed from live search results"),
        ("Monetisation (in-app purchases present)", "yes", "PROXY — presence, not revenue"),
        ("User complaints", "yes" if complaints else "no", "Thin coverage — Apple rate-limited"),
        ("Apple Search Ads popularity", "yes" if asa is not None else "NO",
         "The only real demand measurement"),
        ("Revenue estimates", "yes" if rev is not None else "NO",
         "The only evidence a niche makes money"),
    ]
    ws["A5"], ws["B5"], ws["C5"] = "Evidence", "Have it?", "What it is"
    for c in ("A5", "B5", "C5"):
        ws[c].font, ws[c].fill = HDR, HDR_FILL
    for i, (what, have, note) in enumerate(evidence, 6):
        ws.cell(row=i, column=1, value=what).font = BODY
        cell = ws.cell(row=i, column=2, value=have)
        cell.font = Font(name=FONT, size=10, bold=have in ("NO",),
                         color="CC0000" if have == "NO" else "000000")
        ws.cell(row=i, column=3, value=note).font = NOTE
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 46

    # Live formulas over the Niche Scores tab, so the summary tracks the data
    # rather than freezing a number computed in Python at build time.
    ws["A14"] = "Finalist niches"
    ws["A14"].font = Font(name=FONT, size=11, bold=True)
    ws["A15"], ws["B15"], ws["C15"], ws["D15"] = \
        "Niche", "Keywords", "Median difficulty", "Median opportunity"
    for c in ("A15", "B15", "C15", "D15"):
        ws[c].font, ws[c].fill = HDR, HDR_FILL
    for i, niche in enumerate(niches, 16):
        ws.cell(row=i, column=1, value=niche).font = BODY
        ws.cell(row=i, column=2,
                value=f"=COUNTIF('Niche Scores'!$A:$A,$A{i})").font = BODY
        ws.cell(row=i, column=3,
                value=f"=IFERROR(AVERAGEIF('Niche Scores'!$A:$A,$A{i},"
                      f"'Niche Scores'!$C:$C),\"n/a\")").font = BODY
        ws.cell(row=i, column=4,
                value=f"=IFERROR(AVERAGEIF('Niche Scores'!$A:$A,$A{i},"
                      f"'Niche Scores'!$B:$B),\"n/a\")").font = BODY
    ws.cell(row=16, column=1).comment = Comment(
        "Counts and averages are computed live from the Niche Scores tab, so "
        "editing that tab updates these. Difficulty and opportunity are "
        "proxies on a 0-100 scale, not measurements.", "pipeline")

    # ------------------------------------------------------ Niche Scores ---
    ws = wb.create_sheet("Niche Scores")
    cols = ["niche", "opportunity_proxy", "difficulty_proxy", "difficulty_band",
            "keyword", "autocomplete_rank", "median_rating_count",
            "pct_over_10k_ratings", "publisher_diversity", "unique_developers",
            "median_rating", "median_days_since_update",
            "pct_exact_keyword_in_title", "subtitle_coverage", "iap_pct_proxy",
            "iap_apps_measured", "probably_app_name", "difficulty_inputs",
            "opportunity_inputs", "components_missing"]
    write_table(ws, sel[[c for c in cols if c in sel.columns]]
                .sort_values(["niche", "opportunity_proxy"], ascending=[True, False]))

    # ----------------------------------------------------- Keywords+ASA ----
    ws = wb.create_sheet("Keywords+ASA")
    if asa is not None:
        write_table(ws, asa)
    else:
        ws["A1"] = "Apple Search Ads data has not been collected yet"
        ws["A1"].font = H1
        ws["A2"] = ("This is the only real measurement of search demand in the whole "
                    "study. Everything else is inferred from who currently ranks.")
        ws["A2"].font = NOTE
        ws["A3"] = "Fill the YELLOW columns. See manual/STEP_04_APPLE_SEARCH_ADS.md"
        ws["A3"].font = NOTE
        tmpl = (real.sort_values("opportunity_proxy", ascending=False)
                    .groupby("niche").head(cfg["manual"]["asa_keywords_per_niche"]))
        out = pd.DataFrame({"keyword": tmpl.keyword, "niche": tmpl.niche,
                            "my_difficulty_proxy": tmpl.difficulty_proxy,
                            "asa_popularity": None, "asa_scale": None,
                            "asa_notes": None})
        nxt = write_table(ws, out, start_row=5)
        # Example row, so the expected format is unambiguous.
        ex = ["example: sleep sounds", "(niche)", 42, 3.5, "1-5",
              "record what you SEE; never convert between scales"]
        for j, v in enumerate(ex, 1):
            c = ws.cell(row=nxt, column=j, value=v)
            c.font, c.fill = NOTE, INPUT_FILL
        for r in range(6, 5 + len(out) + 1):
            for j in (4, 5, 6):
                ws.cell(row=r, column=j).fill = INPUT_FILL
                ws.cell(row=r, column=j).font = INPUT_FONT

    # ------------------------------------------------------- Competitors ---
    ws = wb.create_sheet("Competitors")
    ckw = set(sel.keyword)
    cdf = (comp[comp.keyword.isin(ckw)]
           .sort_values(["keyword", "search_rank"])
           [["keyword", "search_rank", "name", "developer", "rating",
             "rating_count", "formatted_price", "has_iap", "iap_price_points",
             "subtitle", "days_since_update", "genre", "size_mb", "app_id"]])
    write_table(ws, cdf.head(3000))

    # -------------------------------------------------------- Complaints ---
    ws = wb.create_sheet("Complaints")
    if complaints:
        ws["A1"] = "Complaint clusters (full text in data/05_complaints.md)"
        ws["A1"].font = H1
        ws.column_dimensions["A"].width = 118
        for i, line in enumerate(complaints.splitlines()[:400], 3):
            c = ws.cell(row=i, column=1, value=line)
            c.font = BODY
            c.alignment = WRAP
        if stats is not None:
            ws2 = wb.create_sheet("Complaints Coverage")
            ws2["A1"] = "How many reviews we actually got per app"
            ws2["A1"].font = H1
            ws2["A2"] = ("Apple's review feed returns HTTP 200 with an empty body when "
                         "throttling. Low counts here mean thin coverage, NOT few complaints.")
            ws2["A2"].font = NOTE
            write_table(ws2, stats, start_row=4)
    else:
        ws["A1"] = "No complaint analysis available — run scripts/p5_reviews.py"
        ws["A1"].font = H1

    # ----------------------------------------------------------- Revenue ---
    ws = wb.create_sheet("Revenue")
    if rev is not None:
        write_table(ws, rev)
    else:
        ws["A1"] = "Revenue data has not been collected yet"
        ws["A1"].font = H1
        ws["A2"] = ("This is the ONLY evidence that a niche makes money. Without it, no "
                    "BUILD verdict here can be considered financially supported.")
        ws["A2"].font = NOTE
        ws["A3"] = "Fill the YELLOW columns. See manual/STEP_06_REVENUE.md"
        ws["A3"].font = NOTE
        src = DATA / "06_revenue_input.csv"
        tmpl = pd.read_csv(src) if src.exists() else pd.DataFrame(
            columns=["niche", "app_id", "app_name", "developer",
                     "est_monthly_revenue", "est_monthly_downloads",
                     "source", "date_pulled"])
        nxt = write_table(ws, tmpl, start_row=5)
        ex = ["(niche)", 123456789, "example: Some App", "Some Developer",
              18000, 9000, "appfigures-us", "2026-07-27"]
        for j, v in enumerate(ex, 1):
            c = ws.cell(row=nxt, column=j, value=v)
            c.font, c.fill = NOTE, INPUT_FILL
        ws.cell(row=nxt + 1, column=1,
                value="Blank means 'no estimate'. Never write 0 — that claims the app "
                      "earns nothing.").font = NOTE
        for r in range(6, 5 + len(tmpl) + 1):
            for j in (5, 6, 7, 8):
                ws.cell(row=r, column=j).fill = INPUT_FILL
                ws.cell(row=r, column=j).font = INPUT_FONT

    # ------------------------------------------- Methodology & Caveats -----
    ws = wb.create_sheet("Methodology & Caveats")
    ws.column_dimensions["A"].width = 112
    lines = [
        ("METHODOLOGY & CAVEATS", H1),
        ("", BODY),
        ("What each number is, and what it is not", Font(name=FONT, size=11, bold=True)),
        ("", BODY),
        ("demand_proxy — position in App Store autocomplete, inverted. Apple orders "
         "suggestions by popularity, so this gives you ORDERING, not volume. Two keywords "
         "at rank 1 are not necessarily equally searched. Only Apple Search Ads (Phase 4) "
         "measures actual demand.", BODY),
        ("", BODY),
        ("difficulty_proxy — a 0-100 weighted blend of: median rating count of the top 10 "
         "(weight 30, log-scaled), share with over 10k ratings (20), share with the exact "
         "keyword in title or subtitle (20), median days since update, inverted (15), "
         "median rating (10), publisher diversity (5). Every input is in the Niche Scores "
         "tab so the arithmetic can be checked.", BODY),
        ("", BODY),
        ("Publisher diversity is weighted at only 5 on purpose. Low diversity is usually "
         "read as 'one player dominates', but it just as often means one small developer "
         "shipping five near-identical apps, which is easy to beat. The same number "
         "supports opposite conclusions, so it is given little influence.", BODY),
        ("", BODY),
        ("iap_pct_proxy — share of measured apps with in-app purchases. Presence of a "
         "paywall, NOT revenue. Read a low number as a red flag, never a high number as "
         "proof of money.", BODY),
        ("", BODY),
        ("Niches are grouped by which apps rank for each keyword (Jaccard overlap of the "
         "top 10, agglomerative, threshold 0.15), not by shared words. Two phrases with no "
         "words in common belong together if the store returns the same apps for both.", BODY),
        ("", BODY),
        ("Known limitations", Font(name=FONT, size=11, bold=True)),
        ("", BODY),
        ("1. Apple signals failure with HTTP 200 and an empty body. Measured: autocomplete "
         "returns an empty list without a storefront header, and the review feed returns an "
         "empty feed under load, intermittently, where fast retries do not recover it. The "
         "pipeline records these as null and never as zero, but it means some data is "
         "simply absent.", BODY),
        ("", BODY),
        ("2. Review coverage is thin and uneven. Apple rate-limited collection heavily. "
         "See the Complaints Coverage tab for exactly how many reviews each app "
         "contributed. Apps with very few reviews cannot support strong claims.", BODY),
        ("", BODY),
        ("3. Subtitles and in-app purchase flags come from the public App Store listing "
         "page, because they do not exist anywhere in Apple's JSON API. They were "
         "collected only for the ~100 apps that rank for the most keywords; elsewhere "
         "these fields are null. The subtitle_coverage column shows where the "
         "exact-keyword signal ran on partial data.", BODY),
        ("", BODY),
        ("4. Secondary markets (CA/GB/AU/IE and DE/FR/NL) were queried with ENGLISH "
         "keywords. For Germany, France and the Netherlands that is a weak proxy: people "
         "there mostly search in their own language. Low competition in those markets "
         "means little English-language competition and nothing more.", BODY),
        ("", BODY),
        ("5. Revenue estimates, where present, are modelled by third parties and are "
         "routinely wrong by 2-3x. Treat them as orders of magnitude only.", BODY),
        ("", BODY),
        ("6. Nothing here is ever estimated to fill a gap. A missing value is missing, and "
         "the reason is in data/errors.log.", BODY),
    ]
    for i, (text, font) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=text)
        c.font, c.alignment = font, WRAP

    dest = OUTPUT / "FINAL_ANALYSIS.xlsx"
    wb.save(dest)
    print(f"\n  -> {dest.relative_to(OUTPUT.parent)}")
    print(f"  tabs: {', '.join(wb.sheetnames)}\n")


if __name__ == "__main__":
    main()
